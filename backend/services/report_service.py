"""
backend/services/report_service.py - CSV and Excel report generation
"""
from __future__ import annotations

import csv
import io
from typing import Any, Optional

from openpyxl import Workbook

from backend.repositories import analysis_repo, candidate_repo


def generate_ranking_csv(company_id: int, jd_id: Optional[int] = None) -> bytes:
    """Generate a ranking CSV for all candidates."""
    rows, _ = candidate_repo.list_candidates(
        company_id=company_id,
        jd_id=jd_id,
        sort_by="score_desc",
        page_size=1000,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "Rank",
            "Name",
            "Email",
            "Current Title",
            "Experience (yrs)",
            "Overall Score (%)",
            "Skill Match (%)",
            "Semantic Match (%)",
            "Recommendation",
            "Pipeline Stage",
            "GitHub",
            "LinkedIn",
            "Reasoning",
            "Uploaded At",
        ]
    )

    for i, row in enumerate(rows, 1):
        writer.writerow(
            [
                i,
                row.get("name", ""),
                row.get("email", ""),
                row.get("current_title", ""),
                row.get("experience_years", 0),
                round(float(row.get("overall_score", 0) or 0) * 100, 1),
                round(float(row.get("skill_match", 0) or 0) * 100, 1),
                round(float(row.get("semantic_match", 0) or 0) * 100, 1),
                row.get("recommendation", ""),
                row.get("pipeline_stage", ""),
                row.get("github", ""),
                row.get("linkedin", ""),
                row.get("reasoning", ""),
                row.get("created_at", ""),
            ]
        )

    return output.getvalue().encode("utf-8")


def generate_pipeline_csv(company_id: int, jd_id: Optional[int] = None) -> bytes:
    """Generate a pipeline status CSV."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Stage", "Candidate Count"])

    funnel = analysis_repo.get_pipeline_funnel(company_id, jd_id)
    for stage, count in zip(funnel["stages"], funnel["counts"]):
        writer.writerow([stage, count])

    return output.getvalue().encode("utf-8")


def generate_analytics_csv(company_id: int, jd_id: Optional[int] = None) -> bytes:
    """Generate an analytics summary CSV."""
    output = io.StringIO()
    writer = csv.writer(output)

    kpis = analysis_repo.get_kpi_counts(company_id, jd_id)
    writer.writerow(["=== Dashboard KPIs ==="])
    writer.writerow(["Metric", "Value"])
    for key, value in kpis.items():
        writer.writerow([key, value])
    writer.writerow([])

    score_distribution = analysis_repo.get_score_distribution(company_id, jd_id)
    writer.writerow(["=== Score Distribution ==="])
    writer.writerow(["Range", "Count"])
    for bucket, count in zip(score_distribution["buckets"], score_distribution["counts"]):
        writer.writerow([bucket, count])
    writer.writerow([])

    missing_skills = analysis_repo.get_missing_skills_frequency(company_id, jd_id)
    writer.writerow(["=== Most Missing Skills ==="])
    writer.writerow(["Skill", "Frequency"])
    for skill, count in zip(missing_skills.get("skills", []), missing_skills.get("counts", [])):
        writer.writerow([skill, count])

    return output.getvalue().encode("utf-8")


def generate_candidate_report(candidate_db_id: int) -> bytes:
    """Generate a single candidate CSV report."""
    candidate = candidate_repo.get_candidate(candidate_db_id)
    if not candidate:
        return b"Candidate not found."

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Candidate Report"])
    writer.writerow([])

    analysis = candidate.get("analysis") or {}
    writer.writerow(["Field", "Value"])
    writer.writerow(["Name", candidate.get("name", "")])
    writer.writerow(["Email", candidate.get("email", "")])
    writer.writerow(["Phone", candidate.get("phone", "")])
    writer.writerow(["Title", candidate.get("current_title", "")])
    writer.writerow(["Experience (yrs)", candidate.get("experience_years", 0)])
    writer.writerow(["GitHub", candidate.get("github", "")])
    writer.writerow(["LinkedIn", candidate.get("linkedin", "")])
    writer.writerow(["Overall Score", f"{float(analysis.get('overall_score', 0) or 0) * 100:.1f}%"])
    writer.writerow(["Skill Match", f"{float(analysis.get('skill_match', 0) or 0) * 100:.1f}%"])
    writer.writerow(["Semantic Match", f"{float(analysis.get('semantic_match', 0) or 0) * 100:.1f}%"])
    writer.writerow(["Recommendation", analysis.get("recommendation", "")])
    writer.writerow(["Pipeline Stage", candidate.get("pipeline_stage", "")])
    writer.writerow(["AI Summary", analysis.get("ai_summary", "")])
    writer.writerow(["Reasoning", analysis.get("reasoning", "")])
    writer.writerow([])

    writer.writerow(["=== Skills ==="])
    writer.writerow(["Skill", "Proficiency", "Duration (months)"])
    for skill in candidate.get("skills", []):
        writer.writerow([skill.get("name", ""), skill.get("proficiency", ""), skill.get("duration_months", 0)])

    return output.getvalue().encode("utf-8")


def generate_final_report(
    company_id: int,
    jd_id: Optional[int] = None,
    jd_label: str = "All JDs",
    selected_candidate_id: Optional[int] = None,
) -> bytes:
    """Generate a single XLSX workbook with all report sections."""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    kpis = analysis_repo.get_kpi_counts(company_id, jd_id)
    score_dist = analysis_repo.get_score_distribution(company_id, jd_id)
    missing_skills = analysis_repo.get_missing_skills_frequency(company_id, jd_id)
    funnel = analysis_repo.get_pipeline_funnel(company_id, jd_id)
    ranking_rows, _ = candidate_repo.list_candidates(
        company_id=company_id,
        jd_id=jd_id,
        sort_by="score_desc",
        page_size=1000,
    )

    _append_rows(
        summary_sheet,
        [
            ["Final Hiring Report"],
            ["Active Job Description", jd_label],
            ["Candidates Included", len(ranking_rows)],
            [],
            ["Dashboard KPI", "Value"],
            *[[key, value] for key, value in kpis.items()],
            [],
            ["Score Bucket", "Count"],
            *[[bucket, count] for bucket, count in zip(score_dist["buckets"], score_dist["counts"])],
            [],
            ["Pipeline Stage", "Count"],
            *[[stage, count] for stage, count in zip(funnel["stages"], funnel["counts"])],
            [],
            ["Top Missing Skill", "Frequency"],
            *[[skill, count] for skill, count in zip(missing_skills.get("skills", []), missing_skills.get("counts", []))],
        ],
    )

    ranking_sheet = workbook.create_sheet("Candidate Ranking")
    ranking_sheet.append(
        [
            "Rank",
            "Name",
            "Email",
            "Current Title",
            "Experience (yrs)",
            "Overall Score (%)",
            "Skill Match (%)",
            "Semantic Match (%)",
            "Recommendation",
            "Pipeline Stage",
        ]
    )
    for idx, row in enumerate(ranking_rows, 1):
        ranking_sheet.append(
            [
                idx,
                row.get("name", ""),
                row.get("email", ""),
                row.get("current_title", ""),
                float(row.get("experience_years", 0) or 0),
                round(float(row.get("overall_score", 0) or 0) * 100, 1),
                round(float(row.get("skill_match", 0) or 0) * 100, 1),
                round(float(row.get("semantic_match", 0) or 0) * 100, 1),
                row.get("recommendation", ""),
                row.get("pipeline_stage", ""),
            ]
        )

    analytics_sheet = workbook.create_sheet("Analytics")
    _append_rows(
        analytics_sheet,
        [
            ["Metric", "Value"],
            *[[key, value] for key, value in kpis.items()],
            [],
            ["Score Range", "Count"],
            *[[bucket, count] for bucket, count in zip(score_dist["buckets"], score_dist["counts"])],
            [],
            ["Missing Skill", "Frequency"],
            *[[skill, count] for skill, count in zip(missing_skills.get("skills", []), missing_skills.get("counts", []))],
        ],
    )

    pipeline_sheet = workbook.create_sheet("Pipeline")
    pipeline_sheet.append(["Stage", "Candidate Count"])
    for stage, count in zip(funnel["stages"], funnel["counts"]):
        pipeline_sheet.append([stage, count])

    if selected_candidate_id:
        _append_candidate_sheet(workbook.create_sheet("Selected Candidate"), selected_candidate_id)

    for sheet in workbook.worksheets:
        _autosize(sheet)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_candidate_sheet(sheet, candidate_db_id: int) -> None:
    candidate = candidate_repo.get_candidate(candidate_db_id)
    if not candidate:
        sheet.append(["Candidate not found"])
        return

    analysis = candidate.get("analysis") or {}
    _append_rows(
        sheet,
        [
            ["Selected Candidate Report"],
            ["Name", candidate.get("name", "")],
            ["Email", candidate.get("email", "")],
            ["Phone", candidate.get("phone", "")],
            ["Title", candidate.get("current_title", "")],
            ["Experience (yrs)", candidate.get("experience_years", 0)],
            ["Pipeline Stage", candidate.get("pipeline_stage", "")],
            ["Overall Score", f"{float(analysis.get('overall_score', 0) or 0) * 100:.1f}%"],
            ["Skill Match", f"{float(analysis.get('skill_match', 0) or 0) * 100:.1f}%"],
            ["Semantic Match", f"{float(analysis.get('semantic_match', 0) or 0) * 100:.1f}%"],
            ["Recommendation", analysis.get("recommendation", "")],
            ["AI Summary", analysis.get("ai_summary", "")],
            ["Reasoning", analysis.get("reasoning", "")],
            [],
            ["Skill", "Proficiency", "Duration (months)"],
            *[
                [skill.get("name", ""), skill.get("proficiency", ""), skill.get("duration_months", 0)]
                for skill in candidate.get("skills", [])
            ],
        ],
    )


def _append_rows(sheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = max((len(value) for value in values), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 48)
