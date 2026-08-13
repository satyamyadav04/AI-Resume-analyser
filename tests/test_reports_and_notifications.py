from __future__ import annotations

from io import BytesIO

import backend.database.db as db_module
from backend.database.db import get_db, init_db
from backend.repositories import analysis_repo, candidate_repo
from backend.services import jd_service
from backend.services.notification_service import send_interview_schedule_email
from backend.services.report_service import generate_final_report
from openpyxl import load_workbook
import pytest


@pytest.fixture()
def isolated_db(tmp_path, monkeypatch):
    existing = getattr(db_module._local, "conn", None)
    if existing is not None:
        existing.close()
        db_module._local.conn = None

    test_db_path = tmp_path / "ats-test.db"
    monkeypatch.setattr(db_module, "_DB_PATH", str(test_db_path))
    init_db()

    yield

    conn = getattr(db_module._local, "conn", None)
    if conn is not None:
        conn.close()
        db_module._local.conn = None


def _seed_base_records() -> tuple[int, int, int]:
    with get_db() as db:
        company_id = db.execute(
            "INSERT INTO companies (name) VALUES (?)",
            ("Report Test Company",),
        ).lastrowid
        user_id = db.execute(
            """
            INSERT INTO users (company_id, name, email, password_hash, role)
            VALUES (?, ?, ?, ?, ?)
            """,
            (company_id, "Recruiter", "recruiter@example.com", "hash", "hr_manager"),
        ).lastrowid

    jd_id = jd_service.create_job_description(
        int(company_id),
        int(user_id),
        "Data Scientist",
        "Build ML systems",
        "Python, SQL, ML",
        make_active=True,
    )
    return int(company_id), int(user_id), int(jd_id)


def _seed_candidate(company_id: int, jd_id: int) -> int:
    candidate_id = candidate_repo.save_candidate(
        company_id=company_id,
        jd_id=jd_id,
        upload_id=None,
        candidate_dict={
            "candidate_id": "cand-1",
            "name": "Asha Sharma",
            "email": "asha@example.com",
            "experience_years": 4.5,
            "profile": {
                "current_title": "ML Engineer",
                "summary": "Strong backend and ML profile",
                "location": "Bengaluru",
                "github": "github.com/asha",
                "linkedin": "linkedin.com/in/asha",
            },
            "skills": [
                {"name": "Python", "proficiency": "advanced", "duration_months": 36},
                {"name": "SQL", "proficiency": "advanced", "duration_months": 30},
            ],
        },
    )
    analysis_repo.save_analysis(
        candidate_db_id=candidate_id,
        jd_id=jd_id,
        scores={
            "overall_score": 0.82,
            "skill_match": 0.78,
            "semantic_match": 0.81,
            "experience_score": 0.45,
            "education_score": 0.70,
            "project_score": 0.60,
            "ai_summary": "Strong fit for the role.",
            "recommendation": "Strong Fit",
            "reasoning": "High overlap with the active JD.",
            "missing_skills": ["TensorFlow"],
            "strengths": ["Python"],
            "weaknesses": [],
            "status": "CLEAN",
            "rank_position": 1,
        },
    )
    return candidate_id


def test_generate_final_report_contains_expected_sheets(isolated_db):
    company_id, _, jd_id = _seed_base_records()
    candidate_id = _seed_candidate(company_id, jd_id)

    report_bytes = generate_final_report(
        company_id=company_id,
        jd_id=jd_id,
        jd_label="Data Scientist",
        selected_candidate_id=candidate_id,
    )

    workbook = load_workbook(BytesIO(report_bytes))

    assert workbook.sheetnames == [
        "Summary",
        "Candidate Ranking",
        "Analytics",
        "Pipeline",
        "Selected Candidate",
    ]
    assert workbook["Summary"]["A1"].value == "Final Hiring Report"
    assert workbook["Candidate Ranking"]["B2"].value == "Asha Sharma"
    assert workbook["Selected Candidate"]["A1"].value == "Selected Candidate Report"


def test_interview_email_returns_skip_when_smtp_not_configured(monkeypatch):
    monkeypatch.delenv("ATS_SMTP_HOST", raising=False)
    monkeypatch.delenv("ATS_SMTP_USER", raising=False)
    monkeypatch.delenv("ATS_SMTP_FROM_EMAIL", raising=False)

    result = send_interview_schedule_email(
        candidate_email="candidate@example.com",
        candidate_name="Asha Sharma",
        company_name="Report Test Company",
        jd_title="Data Scientist",
        interview_data={"date": "2026-07-10", "time": "11:00:00", "round": "HR Round"},
    )

    assert result["sent"] is False
    assert result["attempted"] is False
    assert "SMTP is not configured" in str(result["message"])
